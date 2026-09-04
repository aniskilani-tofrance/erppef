#!/usr/bin/env python3
"""Génère docs/modeles/modele-apprenants.xlsx depuis le RÉFÉRENTIEL UNIQUE du code.
Usage : npm run modele  (exporte d'abord le référentiel, puis construit le fichier)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

REF = json.load(open("scripts/.referentiels.json"))
OUT = "docs/modeles/modele-apprenants.xlsx"
VERT = "0F4C3A"; PALE = "EAF4EF"

wb = Workbook()
ws = wb.active
ws.title = "Apprenants"

headers = ["Prénom","Nom","Téléphone","Email","Langue","Niveau","Naissance (JJ/MM/AAAA)","Sexe",
           "Adresse","Commune","CP","Situation","QPV","RQTH","Scolarisation","Prescripteur","Quartier (St-Ouen)",
           "Objectif","Besoin exprimé","Canal de contact","Précision (canal)"]
widths  = [14,16,14,24,12,10,20,10,26,18,8,20,8,8,18,20,22,26,36,22,26]

for i, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=VERT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

examples = [
    ["Ahmed","Karimi","0612345678","","dari","A1","12/05/1988","homme","3 rue des Écoles","Saint-Ouen-sur-Seine","93400",REF["activities"][0],"oui","non",REF["education"][2],REF["prescribers"][0],REF["districts"][1],REF["goals"][0],"Parler aux employeurs pendant les entretiens",REF["sources"][7],"Conseillère de l'agence de Saint-Ouen"],
    ["Olena","Kovalenko","","olena@mail.com","ukrainien","","03/11/1995","femme","","Aubervilliers","93300",REF["activities"][1],"","",REF["education"][3],REF["prescribers"][1],"","","",REF["sources"][0],"Sœur d'une ancienne apprenante"],
]
ex_fill = PatternFill("solid", fgColor=PALE)
for r, row in enumerate(examples, start=2):
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.fill = ex_fill
        c.font = Font(italic=True, color="5D6F68")

for col in ["C", "G", "K"]:  # téléphone, naissance, CP : texte
    for row in range(2, 501):
        ws[f"{col}{row}"].number_format = "@"

def dv(values, cols):
    formula = '"' + ",".join(values).replace('"', "'") + '"'
    v = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True,
                       errorTitle="Valeur invalide", error="Choisissez une valeur de la liste (ou laissez vide).")
    ws.add_data_validation(v)
    for col in cols:
        v.add(f"{col}2:{col}500")

dv(REF["levels"], ["F"])
dv(REF["genders"], ["H"])
dv(REF["activities"], ["L"])
dv(["oui", "non"], ["M", "N"])
dv(REF["education"], ["O"])
dv(REF["prescribers"], ["P"])
dv(REF["districts"], ["Q"])
dv(REF["goals"], ["R"])
dv(REF["sources"], ["T"])

notice = wb.create_sheet("Notice")
notice.column_dimensions["A"].width = 100
lines = [
    ("MODÈLE D'IMPORT DES APPRENANTS — ERP ParlerEmploi", True),
    ("Généré automatiquement depuis le référentiel de l'application : les menus déroulants sont IDENTIQUES à ceux de l'ERP.", False),
    ("", False),
    ("1. Remplissez la feuille « Apprenants » : une ligne par personne. Seuls Prénom et Nom sont obligatoires.", False),
    ("2. Supprimez les 2 lignes d'exemple (en vert pâle).", False),
    ("3. Déposez ce fichier dans le dossier Drive partagé « Apprenant ERPPEF » : les nouvelles lignes sont importées chaque nuit (ou bouton « Synchroniser le Drive »).", False),
    ("   Autre voie : ERP → Apprenants → « Importer une liste » → « Choisir un fichier Excel ».", False),
    ("", False),
    ("CONSEILS", True),
    ("• Naissance au format JJ/MM/AAAA (colonne en texte : tapez 12/05/1988 tel quel).", False),
    ("• Niveau : laissez VIDE si inconnu — un test de positionnement sera généré automatiquement.", False),
    ("• Quartier : uniquement pour les résidents de Saint-Ouen.", False),
    ("• La typologie alimente directement vos bilans financeurs : plus c'est rempli, mieux c'est.", False),
    ("• Objectif + Besoin exprimé = l'analyse du besoin Qualiopi (ind. 4), reprise dans le dossier d'entrée PDF.", False),
    ("• Canal de contact = comment la personne NOUS a contactés (bouche-à-oreille, France Travail, réseaux sociaux…) ; Précision = nom du partenaire, page, etc.", False),
    ("• Maximum 200 lignes par import manuel (la synchronisation Drive n'a pas cette limite).", False),
]
for r, (text, bold) in enumerate(lines, start=1):
    c = notice.cell(row=r, column=1, value=text)
    c.font = Font(bold=bold, size=12 if bold else 11, color=VERT if bold else "16211D")
    c.alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print("modèle généré :", OUT)
